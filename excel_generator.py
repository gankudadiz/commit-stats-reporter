# -*- coding: utf-8 -*-
"""
Excel 报告生成模块
"""

import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

import i18n


def generate_excel(all_stats, output_path):
    """
    生成 Excel 报告

    Args:
        all_stats: 所有项目的统计数据
        output_path: 输出文件路径

    Returns:
        输出文件路径
    """
    wb = Workbook()

    # 移除默认sheet
    wb.remove(wb.active)

    # 创建汇总表
    create_summary_sheet(wb, all_stats)

    # 为每个项目创建详细表
    for project_name, stats in all_stats.items():
        create_detail_sheet(wb, project_name, stats)

    # 保存文件
    wb.save(output_path)
    return output_path


def create_summary_sheet(wb, all_stats):
    """
    创建汇总表

    Args:
        wb: Workbook 对象
        all_stats: 所有项目的统计数据
    """
    ws = wb.create_sheet(i18n.t("summary"))

    # 设置列宽
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15

    # 标题样式
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=12, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")

    # 写入标题
    headers = [
        i18n.t("project_name"), i18n.t("commits"), 
        i18n.t("insertions_lines"), i18n.t("deletions_lines"), 
        i18n.t("net_lines"), i18n.t("changed_files")
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = header_alignment

    # 写入数据
    row = 2
    total_commits = 0
    total_insertions = 0
    total_deletions = 0
    total_net = 0
    total_files = 0

    for project_name, stats in all_stats.items():
        project_commits = len(stats)
        project_insertions = sum(s['insertions'] for s in stats)
        project_deletions = sum(s['deletions'] for s in stats)
        project_net = sum(s['net'] for s in stats)
        project_files = sum(s['files'] for s in stats)

        ws.cell(row=row, column=1, value=project_name)
        ws.cell(row=row, column=2, value=project_commits)
        ws.cell(row=row, column=3, value=project_insertions)
        ws.cell(row=row, column=4, value=project_deletions)
        ws.cell(row=row, column=5, value=project_net)
        ws.cell(row=row, column=6, value=project_files)

        total_commits += project_commits
        total_insertions += project_insertions
        total_deletions += project_deletions
        total_net += project_net
        total_files += project_files
        row += 1

    # 写入总计
    ws.cell(row=row, column=1, value=i18n.t("total"))
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=2, value=total_commits)
    ws.cell(row=row, column=2).font = Font(bold=True)
    ws.cell(row=row, column=3, value=total_insertions)
    ws.cell(row=row, column=3).font = Font(bold=True)
    ws.cell(row=row, column=4, value=total_deletions)
    ws.cell(row=row, column=4).font = Font(bold=True)
    ws.cell(row=row, column=5, value=total_net)
    ws.cell(row=row, column=5).font = Font(bold=True)
    ws.cell(row=row, column=6, value=total_files)
    ws.cell(row=row, column=6).font = Font(bold=True)

    # 边框样式
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for r in range(1, row + 1):
        for c in range(1, 7):
            ws.cell(row=r, column=c).border = thin_border
            ws.cell(row=r, column=c).alignment = Alignment(horizontal="center")


def create_detail_sheet(wb, project_name, stats):
    """
    创建项目详细表

    Args:
        wb: Workbook 对象
        project_name: 项目名称
        stats: 项目统计数据
    """
    # Sheet名称限制31个字符
    sheet_name = project_name[:31] if len(project_name) > 31 else project_name

    # 确保sheet名称唯一
    if sheet_name in wb.sheetnames:
        base_name = sheet_name
        counter = 1
        while sheet_name in wb.sheetnames:
            sheet_name = f"{base_name[:28]}_{counter}"
            counter += 1

    ws = wb.create_sheet(sheet_name)

    # 设置列宽
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 50
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 12

    # 标题样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")

    # 写入标题
    if 'hash' in stats[0]:
        headers = [
            i18n.t("commit_hash"), i18n.t("author"), i18n.t("date"), 
            i18n.t("message"), i18n.t("insertions"), i18n.t("deletions"), 
            i18n.t("net"), i18n.t("files")
        ]
    else:
        headers = [
            i18n.t("revision"), i18n.t("author"), i18n.t("date"), 
            i18n.t("message"), i18n.t("insertions"), i18n.t("deletions"), 
            i18n.t("net"), i18n.t("files")
        ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = header_alignment

    # 写入数据
    for row_idx, stat in enumerate(stats, 2):
        if 'hash' in stat:
            ws.cell(row=row_idx, column=1, value=stat['hash'])
        else:
            ws.cell(row=row_idx, column=1, value=stat['revision'])

        ws.cell(row=row_idx, column=2, value=stat['author'])
        ws.cell(row=row_idx, column=3, value=stat['date'])
        ws.cell(row=row_idx, column=4, value=stat['message'])
        ws.cell(row=row_idx, column=5, value=stat['insertions'])
        ws.cell(row=row_idx, column=6, value=stat['deletions'])
        ws.cell(row=row_idx, column=7, value=stat['net'])
        ws.cell(row=row_idx, column=8, value=stat['files'])

    # 边框样式
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for r in range(1, len(stats) + 2):
        for c in range(1, 9):
            ws.cell(row=r, column=c).border = thin_border
            ws.cell(row=r, column=c).alignment = Alignment(horizontal="center")


def get_summary_by_date(all_stats):
    """
    按日期汇总统计

    Args:
        all_stats: 所有项目的统计数据

    Returns:
        按日期汇总的统计字典
    """
    date_summary = {}

    for project_name, stats in all_stats.items():
        for stat in stats:
            date = stat['date']
            if date not in date_summary:
                date_summary[date] = {
                    'commits': 0,
                    'insertions': 0,
                    'deletions': 0,
                    'net': 0,
                    'files': 0
                }

            date_summary[date]['commits'] += 1
            date_summary[date]['insertions'] += stat['insertions']
            date_summary[date]['deletions'] += stat['deletions']
            date_summary[date]['net'] += stat['net']
            date_summary[date]['files'] += stat['files']

    return date_summary
